from django.utils import timezone
from django.contrib.auth import login, authenticate, logout, update_session_auth_hash
from functools import partial
from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm
from django.core.serializers import serialize
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
import openpyxl
from openpyxl.utils import get_column_letter
from .models import Producto, Departamento, Cart, CartItem, CustomUser, EstadoUsuario, EstadoProducto, Solicitud, ProductosSolicitud, Notificacion, RolUsuario
from .forms import ProductoForm, CustomUserForm, SolicitudForm, ProductosSolicitudForm, EditarSolicitudForm
from django.db import models
from django.db.models import Sum
from django.urls import reverse
from django.contrib import messages
import json

# Create your views here.


def home(request):
    return render(request, "core/home.html")

def todos_los_productos(request):
    data = {"list": Producto.objects.all().order_by('idProducto')}
    return render(request, "core/todos_los_productos.html", data)

def reponedor_depdiseño(request):
    # Filtra los productos que pertenecen al departamento de Diseño (ID 2)
    productos_diseño = Producto.objects.filter(departamento_id=2)

    return render(request, 'core/reponedor_depdiseño.html', {'list': productos_diseño})

def reponedor_depinformatica(request):
    # Filtra los productos que pertenecen al departamento de Informática (ID 1)
    productos_informatica = Producto.objects.filter(departamento_id=1)

    return render(request, 'core/reponedor_depinformatica.html', {'list': productos_informatica})


def reponedor_depproductosdisponibles(request):
    # Filtra los productos que pertenecen al departamento de Informática (ID 1)
    productos_disponibles = Producto.objects.filter(estadoProducto=1)

    return render(request, 'core/reponedor_depproductosdisponibles.html', {'list': productos_disponibles})


def reponedor_depproductosdadosdebaja(request):
 # Filtra los productos que pertenecen al departamento de Informática (ID 1)
    productos_dadosdebaja = Producto.objects.filter(estadoProducto=2)

    return render(request, 'core/reponedor_depproductosdadosdebaja.html', {'list': productos_dadosdebaja})



def registro(request):
    if request.method == 'POST':
        form = CustomUserForm(request.POST)
        if form.is_valid():
            if form.cleaned_data['password1'] == form.cleaned_data['password2']:
                custom_user = form.save()
                Cart.objects.create(user=custom_user)

                success_message = 'Registro exitoso. ¡Inicia sesión!'
                messages.success(request, success_message)

                response_data = {'message': success_message}
                return JsonResponse(response_data)
            else:
                error_message = 'Las contraseñas no coinciden. Por favor, inténtalo de nuevo.'
                messages.error(request, error_message)
                response_data = {'message': error_message}
                return JsonResponse(response_data, status=400)
        else:
            error_message = 'Error en el formulario. Por favor, verifica los campos.'
            messages.error(request, error_message)
            response_data = {'message': error_message}
            return JsonResponse(response_data, status=400)
    else:
        form = CustomUserForm()

    return render(request, 'core/registro.html', {'form': form})

@login_required
def cambiar_contraseña(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Actualiza la sesión del usuario
            logout(request)  # Cierra la sesión del usuario
            success_message = 'Contraseña cambiada exitosamente. Por favor, inicia sesión de nuevo.'
            messages.success(request, success_message)
            return JsonResponse({'message': success_message})
        else:
            messages.error(request, 'Por favor, corrige los errores a continuación.')
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'core/cambiar_contraseña.html', {'form': form})




def user_login(request):
    error_message = ''  # Asigna un valor predeterminado

    if request.method == 'POST':
        form = AuthenticationForm(request, request.POST)
        if form.is_valid():
            user = authenticate(request, username=form.cleaned_data['username'], password=form.cleaned_data['password'])
            if user is not None:
                login(request, user)
                success_message = 'Inicio de sesión exitoso.'
                response_data = {'message': success_message}
                return JsonResponse(response_data)

            else:
                error_message = 'Nombre de usuario o contraseña incorrectos.'
                response_data = {'error_message': error_message}
                return JsonResponse(response_data, status=400)
        else:
            error_message = 'Error en el formulario. Por favor, verifica los campos.'
            response_data = {'error_message': error_message}
            return JsonResponse(response_data, status=400)
    else:
        form = AuthenticationForm()

    return render(request, 'core/login.html', {'form': form, 'error_message': error_message})


def custom_logout(request):
    logout(request)
    return redirect('determine_home_page')

@login_required
def carrito_compra(request):
    return render(request, "core/carrito_compra.html")

def producto_ficha(request, id):
    producto = Producto.objects.get(idProducto=id)
    print(f"idProducto: {id}")  # Agrega esta línea para imprimir el valor de idProducto
    data = {"producto": producto}
    return render(request, "core/producto_ficha.html", data)


def es_encargado(user, idRol):
    return user.rol.idRol == idRol

# Validador para encargados
es_encargado_encargado = partial(es_encargado, idRol=3)  # Reemplaza "1" con el ID correcto de "encargado"

@user_passes_test(es_encargado_encargado)
def lista_productos(request):
    data = {"list": Producto.objects.all().order_by('idProducto')}
    return render(request, "core/lista_productos.html", data)

@user_passes_test(es_encargado_encargado)
def producto(request, action, id):
    data = {"mesg": "", "form": ProductoForm(), "action": action, "id": id}

    if action == 'ins':
        if request.method == "POST":
            form = ProductoForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    form.save()
                    data["mesg"] = "¡El producto fue creado correctamente!"
                    return redirect('lista_productos')
                except Exception as e:
                    data["mesg"] = f"¡Hubo un error al crear el producto! {str(e)}"
        data["form"] = ProductoForm()

    elif action == 'upd':
        objeto = get_object_or_404(Producto, idProducto=id)
        if request.method == "POST":
            form = ProductoForm(request.POST, request.FILES, instance=objeto)
            if form.is_valid():
                try:
                    form.save()
                    data["mesg"] = "¡El producto fue actualizado correctamente!"
                    return redirect('lista_productos')
                except Exception as e:
                    data["mesg"] = f"¡Hubo un error al actualizar el producto! {str(e)}"
        data["form"] = ProductoForm(instance=objeto)

    elif action == 'del':
        try:
            producto = Producto.objects.get(idProducto=id)
            producto.delete()
            data["mesg"] = "¡El producto fue eliminado correctamente!"
            return redirect('lista_productos')
        except Producto.DoesNotExist:
            data["mesg"] = "¡El producto ya estaba eliminado!"

    data["list"] = Producto.objects.all().order_by('idProducto')
    return render(request, "core/producto.html", data)

def add_to_cart(request, product_id):
    if request.method == 'POST':
        quantity_str = request.POST.get('quantity', '1')
        
        if not quantity_str.isdigit():
            error_message = 'La cantidad ingresada no es válida.'
            return JsonResponse({'error_message': error_message}, status=400)
        
        quantity = int(quantity_str)
        
        product = Producto.objects.get(pk=product_id)
        user_cart, created = Cart.objects.get_or_create(user=request.user)
        cart_item, created = CartItem.objects.get_or_create(cart=user_cart, product=product)

        # Obtener el valor de stock y stock mínimo del producto
        stock = product.stock
        stock_minimo = product.stockMinimo
        stock_disponible = stock - stock_minimo - cart_item.quantity  # Restar lo que ya está en el carrito

        # Calcular la cantidad máxima que se puede solicitar
        cantidad_maxima = min(stock_disponible, stock - stock_minimo)

        if quantity > cantidad_maxima:
            error_message = f'No puedes solicitar más de {cantidad_maxima} unidades de {product.nombreProducto}.'
            return JsonResponse({'error_message': error_message}, status=400)
        elif quantity < 1:
            error_message = 'La cantidad ingresada no es válida.'
            return JsonResponse({'error_message': error_message}, status=400)
        else:
            cart_item.quantity += quantity
            cart_item.save()
            success_message = f'Se agregaron {quantity} {product.nombreProducto} al carrito.'
            return JsonResponse({'message': success_message})




def remove_from_cart(request, product_id):
    if request.user.is_authenticated:
        try:
            product = Producto.objects.get(idProducto=product_id)
            cart = Cart.objects.get(user=request.user)
            cart_item = CartItem.objects.get(cart=cart, product=product)
            cart_item.delete()
        except (Producto.DoesNotExist, Cart.DoesNotExist, CartItem.DoesNotExist):
            pass  # Puedes manejar errores de la manera que prefieras

    return redirect('carrito_compra')




def clear_cart(request):
    if request.user.is_authenticated:
        cart = Cart.objects.get(user=request.user)
        cart.cartitem_set.all().delete()  # Elimina todos los elementos del carrito
    return redirect('carrito_compra')


def buscar_productos(request):
    query = request.GET.get('q')

    # Realiza la búsqueda en la base de datos
    productos = Producto.objects.filter(
        models.Q(idProducto__icontains=query) | models.Q(nombreProducto__icontains=query)
    )

    return render(request, 'core/todos_los_productos.html', {'list': productos})

    
def carrito_compra(request):
    # Obtener el carrito del usuario actual (esto puede variar dependiendo de tu implementación de autenticación)
    user_cart = Cart.objects.get(user=request.user)
    # Obtener los elementos del carrito para ese carrito
    cart_items = CartItem.objects.filter(cart=user_cart)

    data = {
        "cart_items": cart_items
    }
    return render(request, "core/carrito_compra.html", data)


def producto_ficha(request, id, error_message=None):
    producto = Producto.objects.get(idProducto=id)
    data = {"producto":  producto, "error_message": error_message}
    return render(request, "core/producto_ficha.html", data)


def poblar_bd(request):
    Producto.objects.all().delete()
    Producto.objects.create(nombreProducto='Taladro', esFungible=False, estadoProducto=EstadoProducto.objects.get(idEstadoProducto=1), marca='acme', modelo='0123', stock="9", stockMinimo="6", descripcion='Ta ladrando', imagen="images/taladro.png", departamento=Departamento.objects.get(idDepartamento=1), ubicacionRack="1")
    Producto.objects.create(nombreProducto='Martillo', esFungible=False, estadoProducto=EstadoProducto.objects.get(idEstadoProducto=1), marca='acme', modelo='0223', stock="10", stockMinimo="5", descripcion='El de thor', imagen="images/martillo.png", departamento=Departamento.objects.get(idDepartamento=1), ubicacionRack="1")
    Producto.objects.create(nombreProducto='Pintura', esFungible=False, estadoProducto=EstadoProducto.objects.get(idEstadoProducto=1), marca='acme', modelo='0323', stock="40", stockMinimo="30", descripcion='Pinta', imagen="images/pintura.png", departamento=Departamento.objects.get(idDepartamento=2), ubicacionRack="1")
    Producto.objects.create(nombreProducto='Cable USB', esFungible=False, estadoProducto=EstadoProducto.objects.get(idEstadoProducto=1), marca='acme', modelo='0423', stock="50", stockMinimo="20", descripcion='Conecta USB', imagen="images/usb.png", departamento=Departamento.objects.get(idDepartamento=1), ubicacionRack="1")
    Producto.objects.create(nombreProducto='Cable HDMI', esFungible=False, estadoProducto=EstadoProducto.objects.get(idEstadoProducto=1), marca='acme', modelo='0523', stock="50", stockMinimo="25", descripcion='Conecta HDMI', imagen="images/sinfoto.jpg", departamento=Departamento.objects.get(idDepartamento=1), ubicacionRack="1")
    Producto.objects.create(nombreProducto='Adaptador HDMI', esFungible=True, estadoProducto=EstadoProducto.objects.get(idEstadoProducto=1), marca='acme', modelo='0623', stock="40", stockMinimo="50", descripcion='Adapta HDMI', imagen="images/sinfoto.jpg", departamento=Departamento.objects.get(idDepartamento=1), ubicacionRack="1")
    Producto.objects.create(nombreProducto='Televisor', esFungible=False, estadoProducto=EstadoProducto.objects.get(idEstadoProducto=1), marca='acme', modelo='0723', stock="10", stockMinimo="5", descripcion='Para ver tele', imagen="images/sinfoto.jpg", departamento=Departamento.objects.get(idDepartamento=1), ubicacionRack="1")
    Producto.objects.create(nombreProducto='Pincel', esFungible=True, estadoProducto=EstadoProducto.objects.get(idEstadoProducto=1), marca='acme', modelo='0823', stock="60", stockMinimo="20", descripcion='Para pintar', imagen="images/sinfoto.jpg", departamento=Departamento.objects.get(idDepartamento=2), ubicacionRack="1")
    return redirect(producto, action='ins', id = '-1')




def exportar_productos_a_excel_departamentodiseño(request):
    # Obtiene los datos de la tabla de productos que pertenecen al departamento de Informática (ID 1)
    productos = Producto.objects.filter(departamento_id=2)

    # Crea un nuevo libro de Excel y una hoja de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados de columna
    columnas = [
        "ID de Producto",
        "Nombre del producto",
        "Es fungible",
        "Estado del producto",
        "Marca",
        "Modelo",
        "Stock",
        "Stock Mínimo",
        "Descripción",
        "Rack",
    ]

    for col_num, columna_title in enumerate(columnas, 1):
        col_letra = get_column_letter(col_num)
        ws[f"{col_letra}1"] = columna_title

    # Datos de productos
    for fila, producto in enumerate(productos, 2):
        ws[f"A{fila}"] = producto.idProducto
        ws[f"B{fila}"] = producto.nombreProducto
        ws[f"C{fila}"] = "Sí" if producto.esFungible else "No"
        ws[f"D{fila}"] = producto.estadoProducto.estadoProducto
        ws[f"E{fila}"] = producto.marca
        ws[f"F{fila}"] = producto.modelo
        ws[f"G{fila}"] = producto.stock
        ws[f"H{fila}"] = producto.stockMinimo
        ws[f"I{fila}"] = producto.descripcion
        ws[f"J{fila}"] = producto.ubicacionRack

    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = "attachment; filename=productos_informatica.xlsx"
    wb.save(response)

    return response


def exportar_productos_a_excel_departamentoinformatica(request):
    # Obtiene los datos de la tabla de productos que pertenecen al departamento de Informática (ID 1)
    productos = Producto.objects.filter(departamento_id=1)

    # Crea un nuevo libro de Excel y una hoja de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados de columna
    columnas = [
        "ID de Producto",
        "Nombre del producto",
        "Es fungible",
        "Estado del producto",
        "Marca",
        "Modelo",
        "Stock",
        "Stock Mínimo",
        "Descripción",
        "Rack",
    ]

    for col_num, columna_title in enumerate(columnas, 1):
        col_letra = get_column_letter(col_num)
        ws[f"{col_letra}1"] = columna_title

    # Datos de productos
    for fila, producto in enumerate(productos, 2):
        ws[f"A{fila}"] = producto.idProducto
        ws[f"B{fila}"] = producto.nombreProducto
        ws[f"C{fila}"] = "Sí" if producto.esFungible else "No"
        ws[f"D{fila}"] = producto.estadoProducto.estadoProducto
        ws[f"E{fila}"] = producto.marca
        ws[f"F{fila}"] = producto.modelo
        ws[f"G{fila}"] = producto.stock
        ws[f"H{fila}"] = producto.stockMinimo
        ws[f"I{fila}"] = producto.descripcion
        ws[f"J{fila}"] = producto.ubicacionRack

    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = "attachment; filename=productos_informatica.xlsx"
    wb.save(response)

    return response

def exportar_productos_a_excel_productosdadosdebaja(request):
    # Obtiene los datos de la tabla de productos que pertenecen al departamento de Informática (ID 1)
    productos = Producto.objects.filter(estadoProducto=2)

    # Crea un nuevo libro de Excel y una hoja de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados de columna
    columnas = [
        "ID de Producto",
        "Nombre del producto",
        "Es fungible",
        "Estado del producto",
        "Marca",
        "Modelo",
        "Stock",
        "Stock Mínimo",
        "Descripción",
        "Rack",
    ]

    for col_num, columna_title in enumerate(columnas, 1):
        col_letra = get_column_letter(col_num)
        ws[f"{col_letra}1"] = columna_title

    # Datos de productos
    for fila, producto in enumerate(productos, 2):
        ws[f"A{fila}"] = producto.idProducto
        ws[f"B{fila}"] = producto.nombreProducto
        ws[f"C{fila}"] = "Sí" if producto.esFungible else "No"
        ws[f"D{fila}"] = producto.estadoProducto.estadoProducto
        ws[f"E{fila}"] = producto.marca
        ws[f"F{fila}"] = producto.modelo
        ws[f"G{fila}"] = producto.stock
        ws[f"H{fila}"] = producto.stockMinimo
        ws[f"I{fila}"] = producto.descripcion
        ws[f"J{fila}"] = producto.ubicacionRack

    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = "attachment; filename=productos_informatica.xlsx"
    wb.save(response)

    return response

def exportar_productos_a_excel_productosdisponibles(request):
    # Obtiene los datos de la tabla de productos que pertenecen al departamento de Informática (ID 1)
    productos = Producto.objects.filter(estadoProducto=1)

    # Crea un nuevo libro de Excel y una hoja de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados de columna
    columnas = [
        "ID de Producto",
        "Nombre del producto",
        "Es fungible",
        "Estado del producto",
        "Marca",
        "Modelo",
        "Stock",
        "Stock Mínimo",
        "Descripción",
        "Rack",
    ]

    for col_num, columna_title in enumerate(columnas, 1):
        col_letra = get_column_letter(col_num)
        ws[f"{col_letra}1"] = columna_title

    # Datos de productos
    for fila, producto in enumerate(productos, 2):
        ws[f"A{fila}"] = producto.idProducto
        ws[f"B{fila}"] = producto.nombreProducto
        ws[f"C{fila}"] = "Sí" if producto.esFungible else "No"
        ws[f"D{fila}"] = producto.estadoProducto.estadoProducto
        ws[f"E{fila}"] = producto.marca
        ws[f"F{fila}"] = producto.modelo
        ws[f"G{fila}"] = producto.stock
        ws[f"H{fila}"] = producto.stockMinimo
        ws[f"I{fila}"] = producto.descripcion
        ws[f"J{fila}"] = producto.ubicacionRack

    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = "attachment; filename=productos_informatica.xlsx"
    wb.save(response)

    return response

def exportar_productos_a_excel_mantenedor(request):
    # Obtiene los datos de la tabla de productos
    productos = Producto.objects.all()

    # Crea un nuevo libro de Excel y una hoja de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Encabezados de columna
    columnas = [
        "ID de Producto",
        "Nombre del producto",
        "Es fungible",
        "Estado del producto",
        "Marca",
        "Modelo",
        "Stock",
        "Stock Minimo",
        "Descripción",
        "Departamento",
        "Rack",
    ]

    for col_num, columna_title in enumerate(columnas, 1):
        col_letra = get_column_letter(col_num)
        ws[f"{col_letra}1"] = columna_title

    # Datos de productos
    for fila, producto in enumerate(productos, 2):
        ws[f"A{fila}"] = producto.idProducto
        ws[f"B{fila}"] = producto.nombreProducto
        ws[f"C{fila}"] = "Sí" if producto.esFungible else "No"
        ws[f"D{fila}"] = producto.estadoProducto.estadoProducto
        ws[f"E{fila}"] = producto.marca
        ws[f"F{fila}"] = producto.modelo
        ws[f"G{fila}"] = producto.stock
        ws[f"H{fila}"] = producto.stockMinimo
        ws[f"I{fila}"] = producto.descripcion
        ws[f"J{fila}"] = producto.departamento.idDepartamento
        ws[f"K{fila}"] = producto.ubicacionRack

    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = "attachment; filename=productos.xlsx"
    wb.save(response)

    return response

def determine_home_page(request):
    user = request.user  # Obtén el usuario actualmente autenticado
    if user.is_authenticated:
        if user.rol.idRol == 3:
            return redirect('home')
    return redirect('todos_los_productos')


def solicitud_prestamo(request):
    user = request.user
    carrito_usuario = get_object_or_404(Cart, user=user)
    productos_en_carrito = carrito_usuario.products.all()
    cart_items = CartItem.objects.filter(cart__user=request.user)

    if request.method == 'POST':
        solicitud_form = SolicitudForm(request.POST)
        if solicitud_form.is_valid():
            nueva_solicitud = solicitud_form.save(commit=False)
            nueva_solicitud.usuario = user
            nueva_solicitud.save()

            for producto in productos_en_carrito:
                cantidad = producto.cartitem_set.get(cart=carrito_usuario).quantity
                ProductosSolicitud.objects.create(solicitud=nueva_solicitud, producto=producto, cantidad=cantidad)

            carrito_usuario.products.clear()

            # Crear notificación con el nombre del usuario que envía la solicitud
            mensaje_usuario = f"{user.username} ha enviado una solicitud"

            # Encuentra a todos los usuarios con rol idRol igual a 3
            usuarios_destinatarios = CustomUser.objects.filter(rol__idRol=3)

            # Crear notificaciones para todos los usuarios con el rol correspondiente
            for usuario_destinatario in usuarios_destinatarios:
                Notificacion.objects.create(detalle=mensaje_usuario, usuario=usuario_destinatario, tipo='envio_solicitud')

            # JsonResponse para SweetAlert2
            success_message = 'Solicitud enviada con éxito.'
            response_data = {'message': success_message}
            return JsonResponse(response_data)

    else:
        solicitud_form = SolicitudForm()

    return render(request, 'core/solicitud_prestamo.html', {
        'solicitud_form': solicitud_form,
        'productos': productos_en_carrito,
        'cart_items': [{'product': item.product, 'quantity': item.quantity} for item in cart_items],
    })


def solicitud_exitosa(request):
    return render(request, 'core/solicitud_exitosa.html')

def mis_solicitudes(request):
    solicitudes = Solicitud.objects.filter(usuario=request.user)
    return render(request, 'core/mis_solicitudes.html', {'solicitudes': solicitudes})

@user_passes_test(es_encargado_encargado)
def solicitudes(request):
    user = request.user
    solicitudes = Solicitud.objects.all()

    if request.method == 'POST':
        solicitud_id = request.POST.get('solicitud_id')
        action = request.POST.get('action')

        if action == 'approve':
            solicitud = get_object_or_404(Solicitud, idSolicitud=solicitud_id)
            solicitud.estaAprobado = True
            solicitud.save()
            messages.success(request, 'Solicitud aprobada correctamente.')
        elif action == 'reject':
            solicitud = get_object_or_404(Solicitud, idSolicitud=solicitud_id)
            solicitud.estaAprobado = False
            solicitud.save()
            messages.success(request, 'Solicitud rechazada correctamente.')
        elif action == 'delete':
            Solicitud.objects.filter(idSolicitud=solicitud_id).delete()
            messages.success(request, 'Solicitud eliminada correctamente.')

    return render(request, 'core/solicitudes.html', {'solicitudes': solicitudes})


def actualizar_fecha_devolucion_efectiva(request, solicitud_id):
    solicitud = get_object_or_404(Solicitud, idSolicitud=solicitud_id)

    if request.method == 'POST':
        nueva_fecha_devolucion = request.POST.get('fecha_devolucion_efectiva')
        solicitud.fecha_devolucion_efectiva = nueva_fecha_devolucion
        solicitud.save()

    return redirect('solicitudes')

def editar_solicitud(request, solicitud_id):
    solicitud = Solicitud.objects.get(idSolicitud=solicitud_id)

    if request.method == 'POST':
        form = EditarSolicitudForm(request.POST, instance=solicitud)
        if form.is_valid():
            solicitud = form.save(commit=False)
            esta_aprobado = form.cleaned_data['estaAprobado']
            solicitud.estaAprobado = esta_aprobado
            solicitud.save()

            success_message = 'Cambios guardados correctamente.'
            response_data = {'message': success_message}
            return JsonResponse(response_data)

    else:
        form = EditarSolicitudForm(instance=solicitud)

    return render(request, 'core/editar_solicitud.html', {'form': form, 'solicitud': solicitud})

def notificaciones(request):
    # Obtener todas las notificaciones ordenadas por fecha de notificación de la más reciente a la más antigua
    all_notificaciones = Notificacion.objects.all().order_by('-fecha_notificacion')

    return render(request, 'core/notificaciones.html', {'notificaciones': all_notificaciones})

def obtener_notificaciones(request):
    user = request.user

    # Obtener notificaciones según el tipo de usuario y ordenadas por fecha de notificación de la más reciente a la más antigua
    if user.rol.idRol == 3:
        notificaciones_no_leidas = Notificacion.objects.filter(tipo='envio_solicitud', leido=False).order_by('-fecha_notificacion')
        notificaciones_leidas = Notificacion.objects.filter(tipo='envio_solicitud', leido=True).order_by('-fecha_notificacion')
    else:
        notificaciones_no_leidas = Notificacion.objects.filter(usuario=user, tipo__in=['aprobacion', 'rechazo'], leido=False).order_by('-fecha_notificacion')
        notificaciones_leidas = Notificacion.objects.filter(usuario=user, tipo__in=['aprobacion', 'rechazo'], leido=True).order_by('-fecha_notificacion')

    # Contar notificaciones no leídas
    count_no_leidas = notificaciones_no_leidas.count()

    # Combinar las notificaciones no leídas seguidas por las leídas
    notificaciones = list(notificaciones_no_leidas) + list(notificaciones_leidas)

    jsonData = serialize('json', notificaciones)

    return JsonResponse({'data': jsonData, 'count_no_leidas': count_no_leidas})



def marcar_leida(request, notificacion_id):
  
  notificacion = Notificacion.objects.get(pk=notificacion_id)
  notificacion.leido = True
  notificacion.save()
  
  return JsonResponse({'success': True})